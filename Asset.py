import requests
import xml.etree.ElementTree as ET
from requests.auth import HTTPBasicAuth
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging
import sys
import time

# ============================================================
# CONFIGURATION
# ============================================================
USERNAME   = "your_qualys_username"
PASSWORD   = "your_qualys_password"
CERT_PATH  = "/path/to/your/corporate_cert.pem"  # or False to skip SSL verify
BASE_URL   = "https://qualysapi.qg1.apps.qualys.in"
INPUT_FILE = "assets.xlsx"            # Your input Excel file
ASSET_ID_COL = "Asset ID"            # Exact column name in your Excel
BATCH_SIZE = 50                       # IDs per API call (keep ≤ 100)
RETRY_WAIT = 5                        # Seconds to wait on rate-limit / transient error
# ============================================================

logging.basicConfig(stream=sys.stdout, level=logging.INFO, format="[%(levelname)s] %(message)s")
log = logging.getLogger("QualysInventoryUpdater")

# Fields we attempt to extract — Qualys may surface any of these as the "last inventory" timestamp
CANDIDATE_FIELDS = [
    "lastInventory",       # Global AssetView / CSAM primary field
    "inventoryUpdated",    # alternate naming in some subscriptions
    "lastVulnScan",        # fallback — closest proxy if inventory field absent
    "modified",            # record last-modified timestamp
]


class QualysInventoryUpdater:

    def __init__(self):
        self.base_url    = BASE_URL.rstrip("/")
        self.session     = requests.Session()
        self.auth        = HTTPBasicAuth(USERNAME, PASSWORD)
        self.fo_headers  = {"X-Requested-With": "Python"}
        self.qps_headers = {"Content-Type": "application/xml", "Accept": "application/xml"}

    # ------------------------------------------------------------------
    # Auth
    # ------------------------------------------------------------------
    def login(self):
        r = self.session.post(
            f"{self.base_url}/api/2.0/fo/session/",
            headers=self.fo_headers,
            data={"action": "login", "username": USERNAME, "password": PASSWORD},
            verify=CERT_PATH,
        )
        r.raise_for_status()
        if "QualysSession" not in self.session.cookies:
            raise RuntimeError("Login failed — no QualysSession cookie returned.")
        log.info("Login successful.")

    def logout(self):
        try:
            self.session.post(
                f"{self.base_url}/api/2.0/fo/session/",
                headers=self.fo_headers,
                data={"action": "logout"},
                verify=CERT_PATH,
            )
        except Exception:
            pass
        log.info("Logged out.")

    # ------------------------------------------------------------------
    # Build a batch search request for a list of asset IDs
    # ------------------------------------------------------------------
    def _build_batch_request(self, asset_ids: list) -> bytes:
        """
        Uses the same /qps/rest/2.0/search/am/hostasset endpoint as the
        reference script, but with an IN filter on 'id'.
        Requests only the fields we care about to keep payloads small.
        """
        root = ET.Element("ServiceRequest")

        # Filter: id IN (id1, id2, …)
        filters   = ET.SubElement(root, "filters")
        criteria  = ET.SubElement(filters, "Criteria")
        criteria.set("field", "id")
        criteria.set("operator", "IN")
        criteria.text = ",".join(str(i) for i in asset_ids)

        # Fields to return
        fields  = ET.SubElement(root, "fields")
        host    = ET.SubElement(fields, "HostAsset")
        ET.SubElement(host, "id")
        ET.SubElement(host, "dnsHostName")
        ET.SubElement(host, "netbiosName")
        for f in CANDIDATE_FIELDS:
            ET.SubElement(host, f)

        return ET.tostring(root, encoding="utf-8")

    # ------------------------------------------------------------------
    # Fetch one batch; retry once on transient errors
    # ------------------------------------------------------------------
    def _fetch_batch(self, asset_ids: list) -> list:
        url = f"{self.base_url}/qps/rest/2.0/search/am/hostasset"
        for attempt in range(2):
            r = self.session.post(
                url,
                headers=self.qps_headers,
                auth=self.auth,
                data=self._build_batch_request(asset_ids),
                verify=CERT_PATH,
            )
            if r.status_code == 429 or r.status_code >= 500:
                log.warning(f"HTTP {r.status_code} — waiting {RETRY_WAIT}s before retry...")
                time.sleep(RETRY_WAIT)
                continue
            if r.status_code == 401:
                raise RuntimeError("401 — check API Access + Asset Management permissions.")
            r.raise_for_status()
            return ET.fromstring(r.content).findall(".//HostAsset")
        log.error(f"Batch failed after retries: IDs {asset_ids[:5]}...")
        return []

    # ------------------------------------------------------------------
    # Parse a HostAsset element → dict
    # ------------------------------------------------------------------
    @staticmethod
    def _parse_asset(asset_el: ET.Element) -> dict:
        asset_id = asset_el.findtext("id", default="").strip()
        hostname = (
            asset_el.findtext("dnsHostName", default="").strip()
            or asset_el.findtext("netbiosName", default="").strip()
            or asset_id
        )

        # Walk candidate fields; use the first one that has a value
        inv_updated = ""
        found_field = ""
        for field in CANDIDATE_FIELDS:
            val = (asset_el.findtext(field) or "").strip()
            if val:
                inv_updated = val
                found_field = field
                break

        return {
            "asset_id":    asset_id,
            "hostname":    hostname,
            "inv_updated": inv_updated,
            "source_field": found_field,   # lets you see which XML tag was used
        }

    # ------------------------------------------------------------------
    # Core orchestration
    # ------------------------------------------------------------------
    def fetch_inventory_dates(self, asset_ids: list) -> dict:
        """
        Returns { asset_id_str: {"hostname": …, "inv_updated": …, "source_field": …} }
        """
        results   = {}
        total     = len(asset_ids)
        processed = 0

        for start in range(0, total, BATCH_SIZE):
            batch = asset_ids[start : start + BATCH_SIZE]
            log.info(f"Fetching batch {start // BATCH_SIZE + 1} "
                     f"({len(batch)} IDs, {processed}/{total} done)...")

            assets = self._fetch_batch(batch)
            for el in assets:
                parsed = self._parse_asset(el)
                results[parsed["asset_id"]] = parsed

            # Flag IDs in batch that came back with no data
            for aid in batch:
                if str(aid) not in results:
                    results[str(aid)] = {
                        "hostname": "", "inv_updated": "NOT FOUND", "source_field": ""
                    }

            processed += len(batch)
            time.sleep(0.3)   # be polite to the API

        log.info(f"Finished. {len(results)} asset records retrieved.")
        return results

    # ------------------------------------------------------------------
    # Excel I/O
    # ------------------------------------------------------------------
    def load_asset_ids(self) -> tuple[pd.DataFrame, list]:
        df = pd.read_excel(INPUT_FILE, dtype=str)
        df.columns = df.columns.str.strip()
        if ASSET_ID_COL not in df.columns:
            raise ValueError(
                f"Column '{ASSET_ID_COL}' not found. "
                f"Available columns: {list(df.columns)}"
            )
        ids = df[ASSET_ID_COL].dropna().str.strip().tolist()
        log.info(f"Loaded {len(ids)} asset IDs from '{INPUT_FILE}'.")
        return df, ids

    def update_and_save(self, df: pd.DataFrame, results: dict):
        # Add / overwrite output columns
        def _lookup(row, key):
            aid = str(row[ASSET_ID_COL]).strip()
            return results.get(aid, {}).get(key, "NOT FOUND")

        df["Hostname (Qualys)"]           = df.apply(lambda r: _lookup(r, "hostname"),     axis=1)
        df["INVENTORY_LAST_UPDATED_ON"]   = df.apply(lambda r: _lookup(r, "inv_updated"),  axis=1)
        df["Source XML Field"]            = df.apply(lambda r: _lookup(r, "source_field"),  axis=1)

        output_file = INPUT_FILE.replace(".xlsx", "_updated.xlsx")
        df.to_excel(output_file, index=False, engine="openpyxl")

        # Apply styling
        wb = load_workbook(output_file)
        ws = wb.active
        self._style(ws, df)
        wb.save(output_file)

        log.info(f"Saved updated file: {output_file}")

        # Quick summary
        found    = df[df["INVENTORY_LAST_UPDATED_ON"] != "NOT FOUND"].shape[0]
        missing  = df[df["INVENTORY_LAST_UPDATED_ON"] == "NOT FOUND"].shape[0]
        log.info(f"Summary — Found: {found} | Not found / blank: {missing}")

    # ------------------------------------------------------------------
    # Styling (mirrors reference script)
    # ------------------------------------------------------------------
    @staticmethod
    def _style(ws, df):
        thin        = Side(style="thin", color="CCCCCC")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header row
        for cell in ws[1]:
            cell.fill      = PatternFill("solid", fgColor="1F4E79")
            cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = cell_border

        # Data rows with alternating fill
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = PatternFill("solid", fgColor="EBF3FB") if i % 2 == 0 else None
            for cell in row:
                cell.font      = Font(name="Calibri", size=10)
                cell.border    = cell_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if fill:
                    cell.fill = fill

        # Auto column widths
        for idx, col in enumerate(df.columns, 1):
            max_len = max(
                df.iloc[:, idx - 1].astype(str).map(len).max() if not df.empty else 0,
                len(str(col))
            )
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 4, 60)

        ws.freeze_panes    = "A2"
        ws.auto_filter.ref = ws.dimensions

    # ------------------------------------------------------------------
    # Entry point
    # ------------------------------------------------------------------
    def run(self):
        df, asset_ids = self.load_asset_ids()
        self.login()
        try:
            results = self.fetch_inventory_dates(asset_ids)
            self.update_and_save(df, results)
        finally:
            self.logout()


if __name__ == "__main__":
    QualysInventoryUpdater().run()
