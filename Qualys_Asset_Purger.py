"""
qualys_asset_purger.py
======================
Replicates the manual GAV "Search by Asset ID → Purge Assets" workflow
from the terminal.

HOW IT WORKS
────────────
 Step 1 — VERIFY
   Reads Asset IDs from your Excel file and calls the Qualys AM API
   (POST /qps/rest/2.0/search/am/hostasset) to check which IDs actually
   exist on the platform.
   Prints a clear count, e.g.:  ✔  47 / 50 found on Qualys  |  3 not found

 Step 2 — CONFIRM
   Shows a summary table of every asset that will be purged and asks:
       Proceed with purge? [Y/N]:
   Typing anything other than Y (or y) aborts immediately — nothing is deleted.

 Step 3 — PURGE
   Calls POST /qps/rest/2.0/delete/am/hostasset in batches.
   This is the same API Qualys GAV calls internally when you click
   "Purge Assets" from the Inventory page.  It removes the asset record
   and frees the license.
   Prints live per-batch progress and a final summary.

PERMISSIONS REQUIRED
────────────────────
  • API Access
  • Asset Management  (for the QPS search + delete endpoints)

DEPENDENCIES
────────────
  pip install requests pandas openpyxl
"""

import sys
import time
import logging
import xml.etree.ElementTree as ET
from datetime import datetime

import requests
from requests.auth import HTTPBasicAuth
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURATION  ←  edit these before running
# ──────────────────────────────────────────────────────────────────────────────
USERNAME     = "your_qualys_username"
PASSWORD     = "your_qualys_password"

# Your Qualys API server URL  (e.g. https://qualysapi.qg1.apps.qualys.in)
# Find it: Qualys UI → Help → About → Security Operations Center
BASE_URL     = "https://qualysapi.qg1.apps.qualys.in"

# SSL certificate bundle.  Set to False only in a trusted test environment.
CERT_PATH    = "/path/to/your/corporate_cert.pem"

# Input Excel file — must have a column whose name matches ASSET_ID_COL
INPUT_FILE   = "assets.xlsx"
ASSET_ID_COL = "Asset ID"          # exact column header in your Excel

# How many IDs to send per API call (keep ≤ 100; 50 is a safe default)
BATCH_SIZE   = 50

# Seconds to wait before retrying after a 429 / 5xx response
RETRY_WAIT   = 5
# ──────────────────────────────────────────────────────────────────────────────


# ── Logging (stdout only — clean terminal output) ─────────────────────────────
logging.basicConfig(
    stream=sys.stdout,
    level=logging.INFO,
    format="%(message)s",
)
log = logging.getLogger("qualys_purger")

LINE  = "─" * 65
DLINE = "═" * 65


# ──────────────────────────────────────────────────────────────────────────────
class QualysAssetPurger:

    def __init__(self):
        self.base_url    = BASE_URL.rstrip("/")
        self.session     = requests.Session()
        self.auth        = HTTPBasicAuth(USERNAME, PASSWORD)
        self.fo_headers  = {"X-Requested-With": "QualysAssetPurger"}
        self.qps_headers = {
            "Content-Type": "application/xml",
            "Accept":       "application/xml",
        }

    # ── Session ──────────────────────────────────────────────────────────────
    def login(self):
        r = self.session.post(
            f"{self.base_url}/api/2.0/fo/session/",
            headers=self.fo_headers,
            data={"action": "login", "username": USERNAME, "password": PASSWORD},
            verify=CERT_PATH,
        )
        r.raise_for_status()
        if "QualysSession" not in self.session.cookies:
            raise RuntimeError(
                "Login failed — no session cookie returned.\n"
                "Check USERNAME, PASSWORD and BASE_URL."
            )
        log.info(f"  ✔  Logged in to {self.base_url}")

    def logout(self):
        try:
            self.session.post(
                f"{self.base_url}/api/2.0/fo/session/",
                headers=self.fo_headers,
                data={"action": "logout"},
                verify=CERT_PATH,
            )
        except Exception:
            pass
        log.info("  ✔  Logged out")

    # ── Step 1 helpers: verify ────────────────────────────────────────────────
    def _build_search_xml(self, ids: list) -> bytes:
        root = ET.Element("ServiceRequest")
        crit = ET.SubElement(ET.SubElement(root, "filters"), "Criteria")
        crit.set("field", "id")
        crit.set("operator", "IN")
        crit.text = ",".join(str(i) for i in ids)
        # Only pull the fields we need so the payload stays small
        host = ET.SubElement(ET.SubElement(root, "fields"), "HostAsset")
        for tag in ("id", "dnsHostName", "netbiosName", "address"):
            ET.SubElement(host, tag)
        return ET.tostring(root, encoding="utf-8")

    def _post_qps(self, endpoint: str, body: bytes) -> requests.Response:
        """POST to a QPS endpoint with one automatic retry on 429/5xx."""
        url = f"{self.base_url}{endpoint}"
        for attempt in range(2):
            r = self.session.post(
                url,
                headers=self.qps_headers,
                auth=self.auth,
                data=body,
                verify=CERT_PATH,
            )
            if r.status_code in (429,) or r.status_code >= 500:
                log.info(f"  ⚠  HTTP {r.status_code} — waiting {RETRY_WAIT}s and retrying…")
                time.sleep(RETRY_WAIT)
                continue
            if r.status_code == 401:
                raise RuntimeError(
                    "HTTP 401 — check that your account has "
                    "API Access + Asset Management permissions."
                )
            r.raise_for_status()
            return r
        raise RuntimeError(f"Request to {endpoint} failed after 2 attempts.")

    def verify(self, ids: list) -> dict:
        """
        Query Qualys for each ID in batches.
        Returns { "str_id": {"name": str, "address": str, "found": bool} }
        """
        log.info(f"\n  Checking {len(ids)} asset ID(s) against Qualys…\n")
        results   = {}
        total     = len(ids)

        for start in range(0, total, BATCH_SIZE):
            batch = ids[start: start + BATCH_SIZE]
            end   = min(start + BATCH_SIZE, total)
            log.info(f"  Verifying [{start+1}–{end} of {total}]…")

            r        = self._post_qps("/qps/rest/2.0/search/am/hostasset",
                                      self._build_search_xml(batch))
            elements = ET.fromstring(r.content).findall(".//HostAsset")

            for el in elements:
                aid  = el.findtext("id", default="").strip()
                name = (
                    el.findtext("dnsHostName", default="").strip()
                    or el.findtext("netbiosName",  default="").strip()
                    or el.findtext("address",       default="").strip()
                    or "(no name)"
                )
                addr = el.findtext("address", default="").strip()
                results[aid] = {"name": name, "address": addr, "found": True}

            # IDs with no data back → not found
            for aid in batch:
                if str(aid) not in results:
                    results[str(aid)] = {"name": "", "address": "", "found": False}

            time.sleep(0.3)   # polite gap

        return results

    # ── Step 3 helper: purge ──────────────────────────────────────────────────
    def _build_delete_xml(self, ids: list) -> bytes:
        """
        Builds the ServiceRequest body for DELETE /qps/rest/2.0/delete/am/hostasset
        Filters by id IN (id1,id2,…) — same pattern the platform uses internally.
        """
        root = ET.Element("ServiceRequest")
        crit = ET.SubElement(ET.SubElement(root, "filters"), "Criteria")
        crit.set("field", "id")
        crit.set("operator", "IN")
        crit.text = ",".join(str(i) for i in ids)
        return ET.tostring(root, encoding="utf-8")

    def purge_batch(self, ids: list) -> tuple[list, list]:
        """
        Sends one delete request for a list of IDs.
        Returns (purged_ids, failed_ids).
        """
        r    = self._post_qps("/qps/rest/2.0/delete/am/hostasset",
                               self._build_delete_xml(ids))
        root = ET.fromstring(r.content)
        code = root.findtext("responseCode", default="").strip()

        if code == "SUCCESS":
            purged = [el.text.strip()
                      for el in root.findall(".//HostAsset/id") if el.text]
            # Any ID in the batch that wasn't echoed back = failed
            purged_set = set(purged)
            failed     = [str(i) for i in ids if str(i) not in purged_set]
            return purged, failed
        else:
            msg = root.findtext(".//errorMessage", default="unknown error")
            log.info(f"  ✘  API returned {code}: {msg}")
            return [], [str(i) for i in ids]

    # ── Main flow ─────────────────────────────────────────────────────────────
    def run(self):
        # ── Load Excel ────────────────────────────────────────────────────────
        df = pd.read_excel(INPUT_FILE, dtype=str)
        df.columns = df.columns.str.strip()
        if ASSET_ID_COL not in df.columns:
            raise ValueError(
                f"Column '{ASSET_ID_COL}' not found in {INPUT_FILE}.\n"
                f"Available columns: {list(df.columns)}"
            )
        raw_ids = df[ASSET_ID_COL].dropna().str.strip().tolist()
        if not raw_ids:
            log.info("No asset IDs found in the file. Exiting.")
            return

        log.info(DLINE)
        log.info("  Qualys GAV Asset Purger")
        log.info(DLINE)
        log.info(f"  Input file  : {INPUT_FILE}")
        log.info(f"  Asset IDs   : {len(raw_ids)} loaded")
        log.info(f"  Batch size  : {BATCH_SIZE}")
        log.info(LINE)

        self.login()

        try:
            # ── STEP 1 — VERIFY ───────────────────────────────────────────────
            verified = self.verify(raw_ids)

            found_ids   = [aid for aid, v in verified.items() if v["found"]]
            missing_ids = [aid for aid, v in verified.items() if not v["found"]]

            log.info("")
            log.info(LINE)
            log.info("  VERIFICATION RESULT")
            log.info(LINE)
            log.info(
                f"  ✔  {len(found_ids)} / {len(raw_ids)} asset(s) "
                f"found on Qualys"
            )
            if missing_ids:
                log.info(
                    f"  ✘  {len(missing_ids)} ID(s) NOT found (will be skipped): "
                    f"{missing_ids}"
                )

            if not found_ids:
                log.info("\n  Nothing to purge — all IDs were not found. Exiting.")
                return

            # ── STEP 2 — CONFIRM ──────────────────────────────────────────────
            log.info("")
            log.info(LINE)
            log.info("  ASSETS QUEUED FOR PURGE")
            log.info(LINE)
            log.info(f"  {'Asset ID':<15} {'Name / Hostname':<40} {'IP Address'}")
            log.info(f"  {'─'*15} {'─'*40} {'─'*15}")
            for aid in found_ids:
                info = verified[aid]
                log.info(
                    f"  {aid:<15} {info['name'][:40]:<40} {info['address']}"
                )
            log.info(LINE)
            log.info(
                f"  Total to purge  : {len(found_ids)}"
                + (f"  |  Skipped (not found): {len(missing_ids)}"
                   if missing_ids else "")
            )
            log.info("")
            log.info("  ⚠   Purging is PERMANENT and cannot be undone.")

            try:
                answer = input("  Proceed with purge? [Y/N]: ").strip().lower()
            except (EOFError, KeyboardInterrupt):
                answer = "n"

            if answer != "y":
                log.info("\n  Aborted. No assets were purged.")
                return

            # ── STEP 3 — PURGE ────────────────────────────────────────────────
            log.info("")
            log.info(LINE)
            log.info("  PURGING ASSETS")
            log.info(LINE)

            total_purged  = []
            total_failed  = []
            total         = len(found_ids)

            for start in range(0, total, BATCH_SIZE):
                batch     = found_ids[start: start + BATCH_SIZE]
                batch_num = start // BATCH_SIZE + 1
                end       = min(start + BATCH_SIZE, total)

                log.info(
                    f"  Batch {batch_num}  [{start+1}–{end} of {total}]  "
                    f"purging {len(batch)} asset(s)…"
                )
                purged, failed = self.purge_batch(batch)
                total_purged.extend(purged)
                total_failed.extend(failed)

                log.info(
                    f"           ✔ purged: {len(purged)}"
                    + (f"   ✘ errors: {len(failed)}  {failed}" if failed else "")
                )
                time.sleep(0.5)

            # ── FINAL SUMMARY ─────────────────────────────────────────────────
            log.info("")
            log.info(DLINE)
            log.info("  PURGE COMPLETE")
            log.info(DLINE)
            log.info(f"  Total input IDs        : {len(raw_ids)}")
            log.info(f"  Found on Qualys        : {len(found_ids)}")
            log.info(f"  Successfully purged    : {len(total_purged)}")
            if total_failed:
                log.info(f"  Failed (API error)     : {len(total_failed)}  → {total_failed}")
            if missing_ids:
                log.info(f"  Not found (skipped)    : {len(missing_ids)}")
            log.info(DLINE)

            # ── Save result Excel ─────────────────────────────────────────────
            purged_set  = set(total_purged)
            failed_set  = set(total_failed)
            missing_set = set(missing_ids)

            def _status(row):
                aid = str(row[ASSET_ID_COL]).strip()
                if aid in purged_set:  return "PURGED"
                if aid in failed_set:  return "ERROR"
                if aid in missing_set: return "NOT FOUND"
                return "NOT ATTEMPTED"

            def _name(row):
                aid = str(row[ASSET_ID_COL]).strip()
                return verified.get(aid, {}).get("name", "")

            df["Asset Name (Qualys)"] = df.apply(_name,   axis=1)
            df["Purge Status"]        = df.apply(_status, axis=1)

            ts          = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            output_file = INPUT_FILE.replace(".xlsx", f"_purge_result_{ts}.xlsx")
            df.to_excel(output_file, index=False, engine="openpyxl")
            _apply_styles(output_file, df)
            log.info(f"\n  Results saved → {output_file}\n")

        finally:
            self.logout()


# ── Excel styling ─────────────────────────────────────────────────────────────
_STATUS_COLOR = {
    "PURGED":        "C6EFCE",   # green
    "ERROR":         "FFC7CE",   # red
    "NOT FOUND":     "FFEB9C",   # amber
    "NOT ATTEMPTED": "D9D9D9",   # grey
}

def _apply_styles(filepath: str, df: pd.DataFrame):
    wb = load_workbook(filepath)
    ws = wb.active
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    for cell in ws[1]:
        cell.fill      = PatternFill("solid", fgColor="1F4E79")
        cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border

    # Find "Purge Status" column index (1-based)
    status_col = next(
        (i for i, col in enumerate(df.columns, 1) if col == "Purge Status"),
        None
    )

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        alt_fill = PatternFill("solid", fgColor="EBF3FB") if row_idx % 2 == 0 else None
        for cell in row:
            cell.font      = Font(name="Calibri", size=10)
            cell.border    = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if alt_fill:
                cell.fill = alt_fill
        if status_col:
            sc    = ws.cell(row=row_idx, column=status_col)
            color = _STATUS_COLOR.get(str(sc.value or ""))
            if color:
                sc.fill = PatternFill("solid", fgColor=color)

    # Auto column widths
    for idx, col in enumerate(df.columns, 1):
        col_vals  = df.iloc[:, idx-1].astype(str)
        max_width = max(col_vals.map(len).max() if not df.empty else 0, len(str(col)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_width + 4, 55)

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    QualysAssetPurger().run()
