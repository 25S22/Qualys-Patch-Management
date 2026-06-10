"""
create_test_data.py
───────────────────
Run this once to generate sample input files that match what the Qualys script expects.

Usage:
    python create_test_data.py

Outputs (written to ./test_data/):
    aws_instances_sheet1.xlsx   — Excel file with EC2 instance data
    aws_instances_sheet2.csv    — CSV file with more instances
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_DIR = "./test_data"


def create_excel_file():
    """
    Creates aws_instances_sheet1.xlsx
    Required columns: 'Resource ID' and 'State'
    These match the id_column and status_column parameters in get_dead_instance_dicts()
    """
    data = {
        "Resource ID": [
            "i-0abc1234567890001",  # stopped  → will be picked up
            "i-0abc1234567890002",  # running  → ignored
            "i-0abc1234567890003",  # terminated → will be picked up
            "i-0abc1234567890004",  # running  → ignored
            "i-0abc1234567890005",  # stopped  → will be picked up
            "i-0abc1234567890006",  # running  → ignored
            "i-0abc1234567890007",  # terminated → will be picked up
            "i-0abc1234567890008",  # running  → ignored
        ],
        "State": [
            "stopped",
            "running",
            "terminated",
            "running",
            "stopped",
            "running",
            "terminated",
            "running",
        ],
        # Optional extra columns – the script only uses Resource ID and State
        "Instance Type": ["t2.micro", "t3.medium", "t2.small", "m5.large",
                          "t2.nano", "c5.xlarge", "t3.micro", "r5.large"],
        "Region":        ["ap-south-1"] * 8,
        "Account ID":    ["123456789012"] * 8,
        "Name":          [f"server-{i:03d}" for i in range(1, 9)],
        "Last Seen":     ["2024-01-15", "2024-03-20", "2023-12-01", "2024-03-22",
                          "2024-02-10", "2024-03-22", "2023-11-05", "2024-03-21"],
    }

    df = pd.DataFrame(data)
    filepath = os.path.join(OUTPUT_DIR, "aws_instances_sheet1.xlsx")
    df.to_excel(filepath, index=False, sheet_name="EC2 Instances")

    # Apply basic formatting so it looks like a real export
    wb = load_workbook(filepath)
    ws = wb.active
    header_fill   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    stopped_fill  = PatternFill("solid", start_color="FFE0B2", end_color="FFE0B2")   # orange tint
    term_fill     = PatternFill("solid", start_color="FFCDD2", end_color="FFCDD2")   # red tint
    running_fill  = PatternFill("solid", start_color="C8E6C9", end_color="C8E6C9")   # green tint

    # Format header row
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Highlight State column by value
    state_col_idx = df.columns.tolist().index("State") + 1  # 1-based
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        state_val = row[state_col_idx - 1].value or ""
        if state_val.lower() == "stopped":
            row[state_col_idx - 1].fill = stopped_fill
        elif state_val.lower() == "terminated":
            row[state_col_idx - 1].fill = term_fill
        else:
            row[state_col_idx - 1].fill = running_fill

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    wb.save(filepath)
    print(f"✅  Created: {filepath}  ({len(df)} rows, {len(df[df['State'].str.contains('stopped|terminated')])} dead)")
    return filepath


def create_csv_file():
    """
    Creates aws_instances_sheet2.csv
    Same columns, different instance IDs — tests that the script reads multiple files.
    """
    data = {
        "Resource ID": [
            "i-0def1234567890009",   # running  → ignored
            "i-0def1234567890010",   # stopped  → picked up
            "i-0def1234567890011",   # terminated → picked up
            "i-0def1234567890012",   # running  → ignored
            "i-0def1234567890013",   # stopped  → picked up
        ],
        "State": [
            "running",
            "stopped",
            "terminated",
            "running",
            "stopped",
        ],
        "Instance Type": ["t2.micro", "t3.small", "m5.large", "c5.xlarge", "t2.nano"],
        "Region":        ["us-east-1"] * 5,
        "Account ID":    ["987654321098"] * 5,
        "Name":          [f"server-{i:03d}" for i in range(9, 14)],
        "Last Seen":     ["2024-03-20", "2024-01-08", "2023-10-15", "2024-03-22", "2024-02-28"],
    }

    df = pd.DataFrame(data)
    filepath = os.path.join(OUTPUT_DIR, "aws_instances_sheet2.csv")
    df.to_csv(filepath, index=False)
    print(f"✅  Created: {filepath}  ({len(df)} rows, {len(df[df['State'].str.contains('stopped|terminated')])} dead)")
    return filepath


def print_summary(excel_path, csv_path):
    print("\n" + "=" * 60)
    print("TEST DATA SUMMARY")
    print("=" * 60)
    for path in [excel_path, csv_path]:
        if path.endswith(".xlsx"):
            df = pd.read_excel(path)
        else:
            df = pd.read_csv(path)

        dead = df[df["State"].str.contains("stopped|terminated", case=False, na=False)]
        print(f"\nFile: {os.path.basename(path)}")
        print(f"  Total rows   : {len(df)}")
        print(f"  Dead (target): {len(dead)}")
        print(f"  Dead IDs     :")
        for _, row in dead.iterrows():
            print(f"    → {row['Resource ID']}  [{row['State']}]")

    print("\n" + "=" * 60)
    print("HOW THESE FILES ARE USED BY qualys_purge.py")
    print("=" * 60)
    print("""
  get_dead_instance_dicts(filepath="./test_data", id_column="Resource ID", status_column="State")
  └─ reads every .xlsx and .csv in the folder
  └─ filters rows where State contains 'stopped' or 'terminated'
  └─ returns a list of dicts: [{"Resource ID": "i-0...", "State": "stopped"}, ...]

  The returned Resource IDs are then matched against Qualys assets
  using the EC2 instance ID field:
    sourceInfo.list.Ec2AssetSourceSimple.instanceId
""")


if __name__ == "__main__":
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    excel_path = create_excel_file()
    csv_path   = create_csv_file()
    print_summary(excel_path, csv_path)
    print(f"\nSet FILEPATH = \"{os.path.abspath(OUTPUT_DIR)}\" in qualys_purge.py to use these files.\n")
