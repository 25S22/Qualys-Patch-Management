import requests
import xml.etree.ElementTree as ET
from requests.auth import HTTPBasicAuth
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import logging
import sys

# ============================================================
# CONFIGURATION
# ============================================================
USERNAME  = "your_qualys_username"
PASSWORD  = "your_qualys_password"
CERT_PATH = "/path/to/your/corporate_cert.pem"  # or False to skip SSL verify
BASE_URL  = "https://qualysapi.qg1.apps.qualys.in"
PAGE_SIZE = 100
# ============================================================

# Qualys flattened path for EC2 Instance ID
EC2_RAW_COL = "sourceInfo.list.Ec2AssetSourceSimple.instanceId"

logging.basicConfig(stream=sys.stdout, level=logging.INFO, format="[%(levelname)s] %(message)s")
log = logging.getLogger("QualysFullAssetExtractor")


class QualysFullAssetExtractor:

    def __init__(self):
        self.base_url    = BASE_URL.rstrip("/")
        self.session     = requests.Session()
        self.auth        = HTTPBasicAuth(USERNAME, PASSWORD)
        self.fo_headers  = {"X-Requested-With": "Python"}
        self.qps_headers = {"Content-Type": "application/xml", "Accept": "application/xml"}

    def login(self):
        r = self.session.post(
            f"{self.base_url}/api/2.0/fo/session/",
            headers=self.fo_headers,
            data={"action": "login", "username": USERNAME, "password": PASSWORD},
            verify=CERT_PATH,
        )
        r.raise_for_status()
        if "QualysSession" not in self.session.cookies:
            raise RuntimeError("Login failed — no QualysSession cookie returned.")
        log.info("Login successful.")

    def logout(self):
        try:
            self.session.post(
                f"{self.base_url}/api/2.0/fo/session/",
                headers=self.fo_headers,
                data={"action": "logout"},
                verify=CERT_PATH,
            )
        except Exception:
            pass
        log.info("Logged out.")

    def _build_request(self, offset: int) -> bytes:
        root  = ET.Element("ServiceRequest")
        prefs = ET.SubElement(root, "preferences")
        ET.SubElement(prefs, "startFromOffset").text = str(offset)
        ET.SubElement(prefs, "limitResults").text    = str(PAGE_SIZE)
        return ET.tostring(root, encoding="utf-8")

    def fetch_all_hosts(self) -> list:
        url, offset, page, results = (
            f"{self.base_url}/qps/rest/2.0/search/am/hostasset",
            1, 1, []
        )
        while True:
            log.info(f"Page {page} (offset {offset})...")
            r = self.session.post(
                url, headers=self.qps_headers, auth=self.auth,
                data=self._build_request(offset), verify=CERT_PATH,
            )
            if r.status_code == 401:
                raise RuntimeError(
                    "401 — enable API Access + Asset Management in Administration > Users."
                )
            r.raise_for_status()
            root   = ET.fromstring(r.content)
            assets = root.findall(".//HostAsset")
            if not assets:
                break
            results.extend(assets)
            log.info(f"  {len(assets)} assets on this page | running total: {len(results)}")
            if root.findtext(".//hasMoreRecords") != "true":
                break
            offset += PAGE_SIZE
            page   += 1
        return results

    def _flatten_element(self, element: ET.Element, prefix: str = "") -> dict:
        result   = {}
        children = list(element)

        if not children:
            text = (element.text or "").strip()
            if prefix:
                result[prefix] = text
            return result

        tag_groups = defaultdict(list)
        for child in children:
            tag_groups[child.tag].append(child)

        for tag, siblings in tag_groups.items():
            col_name = f"{prefix}.{tag}" if prefix else tag

            if len(siblings) > 1:
                values = []
                for sib in siblings:
                    leaf_vals = self._flatten_element(sib, "").values()
                    values.append(", ".join(v for v in leaf_vals if v))
                result[col_name] = " | ".join(values)
            else:
                child        = siblings[0]
                grandchildren = list(child)
                if not grandchildren:
                    result[col_name] = (child.text or "").strip()
                else:
                    nested = self._flatten_element(child, col_name)
                    result.update(nested)

        return result

    def build_rows(self, hosts: list) -> pd.DataFrame:
        rows = []
        for host in hosts:
            row = self._flatten_element(host)
            rows.append(row)

        df = pd.DataFrame(rows)

        # ── EC2 Instance ID: promote to a clean top-level column ──────────
        # Non-EC2 assets will simply have a blank value here.
        if EC2_RAW_COL in df.columns:
            df.insert(0, "EC2 Instance ID", df[EC2_RAW_COL])
            log.info(
                f"EC2 Instance ID populated for "
                f"{df['EC2 Instance ID'].astype(bool).sum()} / {len(df)} assets."
            )
        else:
            # Column absent entirely — no EC2 assets in this pull, or field
            # path differs; check _flatten_element output in the column summary.
            df.insert(0, "EC2 Instance ID", "")
            log.warning(
                f"'{EC2_RAW_COL}' not found in response. "
                "EC2 Instance ID column will be blank. "
                "Run with log level DEBUG and inspect the raw XML to confirm the path."
            )

        # Sort columns: identifying ones first
        priority = [
            "EC2 Instance ID", "id", "dnsHostName", "netbiosName", "fqdn",
            "address", "os", "operatingSystem", "trackingMethod", "type",
            "created", "modified", "lastVulnScan", "lastSystemBoot",
            EC2_RAW_COL,          # keep the raw column too for traceability
        ]
        front = [c for c in priority if c in df.columns]
        rest  = [c for c in df.columns if c not in front]
        df    = df[front + sorted(rest)]

        df = df.fillna("")
        log.info(f"DataFrame shape: {df.shape[0]} rows × {df.shape[1]} columns")
        return df

    def export(self, df: pd.DataFrame) -> str:
        filename = "qualys_all_assets_full.xlsx"
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="All Assets", index=False)
            self._style(writer.sheets["All Assets"], df)
        log.info(f"Saved → {filename}")
        return filename

    def _style(self, ws, df):
        thin        = Side(style="thin", color="CCCCCC")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill      = PatternFill("solid", fgColor="1F4E79")
            cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = cell_border

        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = PatternFill("solid", fgColor="EBF3FB") if i % 2 == 0 else None
            for cell in row:
                cell.font      = Font(name="Calibri", size=10)
                cell.border    = cell_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                if fill:
                    cell.fill = fill

        for idx, col in enumerate(df.columns, 1):
            col_vals = df.iloc[:, idx - 1].astype(str).map(len)
            max_len  = max(col_vals.max() if not df.empty else 0, len(str(col)))
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 4, 60)

        ws.freeze_panes    = "A2"
        ws.auto_filter.ref = ws.dimensions

    def run(self):
        self.login()
        try:
            hosts = self.fetch_all_hosts()
            log.info(f"Total HostAsset elements fetched: {len(hosts)}")
            if not hosts:
                log.warning("No hosts returned — check your API permissions and subscription.")
                return

            df = self.build_rows(hosts)

            log.info("\n=== COLUMN SUMMARY ===")
            for col in df.columns:
                non_blank = df[col].astype(bool).sum()
                log.info(f"  {col:<60} filled: {non_blank}/{len(df)}")

            self.export(df)
        finally:
            self.logout()


if __name__ == "__main__":
    QualysFullAssetExtractor().run()
