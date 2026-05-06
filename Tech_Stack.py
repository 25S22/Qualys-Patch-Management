import requests
import xml.etree.ElementTree as ET
from requests.auth import HTTPBasicAuth
import pandas as pd
from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging
import sys

# ============================================================
# CONFIGURATION
# ============================================================
USERNAME  = "your_qualys_username"
PASSWORD  = "your_qualys_password"
CERT_PATH = "/path/to/your/corporate_cert.pem"  # or False to skip SSL verify
BASE_URL  = "https://qualysapi.qg1.apps.qualys.in"
PAGE_SIZE = 100
# ============================================================

logging.basicConfig(stream=sys.stdout, level=logging.INFO, format="[%(levelname)s] %(message)s")
log = logging.getLogger("QualysTechStack")

OS_CATEGORIES = {
    "Windows": ["windows"],
    "Linux":   ["linux", "ubuntu", "debian", "centos", "rhel", "red hat", "suse", "fedora", "amazon"],
    "macOS":   ["mac", "darwin", "os x"],
}

def categorise_os(os_str: str) -> str:
    lower = os_str.lower()
    for label, keywords in OS_CATEGORIES.items():
        if any(k in lower for k in keywords):
            return label
    return "Other"


class QualysTechStackExtractor:

    def __init__(self):
        self.base_url    = BASE_URL.rstrip("/")
        self.session     = requests.Session()
        self.auth        = HTTPBasicAuth(USERNAME, PASSWORD)
        self.fo_headers  = {"X-Requested-With": "Python"}
        self.qps_headers = {"Content-Type": "application/xml", "Accept": "application/xml"}

    def login(self):
        r = self.session.post(
            f"{self.base_url}/api/2.0/fo/session/",
            headers=self.fo_headers,
            data={"action": "login", "username": USERNAME, "password": PASSWORD},
            verify=CERT_PATH,
        )
        r.raise_for_status()
        if "QualysSession" not in self.session.cookies:
            raise RuntimeError("Login failed — no QualysSession cookie returned.")
        log.info("Login successful.")

    def logout(self):
        try:
            self.session.post(
                f"{self.base_url}/api/2.0/fo/session/",
                headers=self.fo_headers,
                data={"action": "logout"},
                verify=CERT_PATH,
            )
        except Exception:
            pass
        log.info("Logged out.")

    def _build_request(self, offset: int) -> bytes:
        # No <filters> block = fetch all hosts
        # Only pulling what we need: id, dns, netbios, os, software name+version
        root  = ET.Element("ServiceRequest")
        prefs = ET.SubElement(root, "preferences")
        ET.SubElement(prefs, "startFromOffset").text = str(offset)
        ET.SubElement(prefs, "limitResults").text    = str(PAGE_SIZE)
        fields  = ET.SubElement(root, "fields")
        host    = ET.SubElement(fields, "HostAsset")
        ET.SubElement(host, "id")
        ET.SubElement(host, "dnsHostName")
        ET.SubElement(host, "netbiosName")
        ET.SubElement(host, "os")
        sw_list = ET.SubElement(host, "HostAssetSoftwareList")
        sw_item = ET.SubElement(sw_list, "HostAssetSoftware")
        ET.SubElement(sw_item, "name")
        ET.SubElement(sw_item, "version")
        return ET.tostring(root, encoding="utf-8")

    def fetch_all_hosts(self) -> list:
        url, offset, page, results = f"{self.base_url}/qps/rest/2.0/search/am/hostasset", 1, 1, []
        while True:
            log.info(f"Page {page} (offset {offset})...")
            r = self.session.post(
                url, headers=self.qps_headers, auth=self.auth,
                data=self._build_request(offset), verify=CERT_PATH,
            )
            if r.status_code == 401:
                raise RuntimeError("401 — enable API Access + Asset Management in Administration > Users.")
            r.raise_for_status()
            root   = ET.fromstring(r.content)
            assets = root.findall(".//HostAsset")
            if not assets:
                break
            results.extend(assets)
            log.info(f"  {len(assets)} assets | total: {len(results)}")
            if root.findtext(".//hasMoreRecords") != "true":
                break
            offset += PAGE_SIZE
            page   += 1
        return results

    def build_rows(self, hosts: list) -> pd.DataFrame:
        """
        Returns one flat DataFrame — one row per (Host x Software x Version).
        All sheets are derived from this single source.
        """
        rows = []
        for host in hosts:
            dns     = host.findtext("dnsHostName", default="")
            netbios = host.findtext("netbiosName",  default="")
            os_str  = host.findtext("os",           default="")
            os_cat  = categorise_os(os_str)
            label   = dns or netbios or host.findtext("id", default="")

            for sw in host.findall(".//HostAssetSoftware"):
                name = (sw.findtext("name")    or "").strip()
                ver  = (sw.findtext("version") or "").strip()
                if not name:
                    continue
                rows.append({
                    "Hostname": label,
                    "OS":       os_str,
                    "OS Group": os_cat,
                    "Software": name,
                    "Version":  ver,
                })
        return pd.DataFrame(rows)

    def make_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        One row per (Software, Version, OS Group).
        Asset Count = number of unique hosts with that exact version on that OS group.
        """
        return (
            df.groupby(["Software", "Version", "OS Group"], sort=True)
            .agg(Asset_Count=("Hostname", "nunique"))
            .reset_index()
            .rename(columns={"Asset_Count": "Asset Count"})
            .sort_values(["Software", "Version", "OS Group"])
            .reset_index(drop=True)
        )

    def export(self, raw_df: pd.DataFrame):
        filename = "qualys_full_techstack.xlsx"
        summary  = self.make_summary(raw_df)

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            # Sheet 1 — full summary across all OS groups
            summary.to_excel(writer, sheet_name="All Software", index=False)
            self._style(writer.sheets["All Software"], summary)

            # One sheet per OS group — same columns, filtered rows
            for grp in sorted(raw_df["OS Group"].unique()):
                grp_df = (
                    summary[summary["OS Group"] == grp]
                    .drop(columns=["OS Group"])
                    .reset_index(drop=True)
                )
                sheet_name = grp[:31]   # Excel sheet name limit
                grp_df.to_excel(writer, sheet_name=sheet_name, index=False)
                self._style(writer.sheets[sheet_name], grp_df)

        log.info(f"Saved: {filename}")
        return filename

    def _style(self, ws, df):
        thin        = Side(style="thin", color="CCCCCC")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill      = PatternFill("solid", fgColor="1F4E79")
            cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = cell_border

        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = PatternFill("solid", fgColor="EBF3FB") if i % 2 == 0 else None
            for cell in row:
                cell.font      = Font(name="Calibri", size=10)
                cell.border    = cell_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if fill:
                    cell.fill = fill

        for idx, col in enumerate(df.columns, 1):
            max_len = max(
                df.iloc[:, idx - 1].astype(str).map(len).max() if not df.empty else 0,
                len(str(col))
            )
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 4, 60)

        ws.freeze_panes  = "A2"
        ws.auto_filter.ref = ws.dimensions

    def run(self):
        self.login()
        try:
            hosts  = self.fetch_all_hosts()
            log.info(f"Total hosts: {len(hosts)}")
            if not hosts:
                log.warning("No hosts returned.")
                return

            raw_df = self.build_rows(hosts)
            log.info(f"Unique software: {raw_df['Software'].nunique()}")
            log.info(f"(Software, Version) combinations: {raw_df.groupby(['Software','Version']).ngroups}")

            self.export(raw_df)

            print("\n=== TOP 20 BY ASSET COUNT (ALL OS) ===")
            print(
                raw_df.groupby("Software")["Hostname"].nunique()
                .reset_index().rename(columns={"Hostname": "Asset Count"})
                .sort_values("Asset Count", ascending=False)
                .head(20).to_string(index=False)
            )
        finally:
            self.logout()


if __name__ == "__main__":
    QualysTechStackExtractor().run()
