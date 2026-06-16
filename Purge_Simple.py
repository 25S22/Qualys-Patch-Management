import time
import os
import sys
import math
import logging
import xml.etree.ElementTree as ET
from datetime import datetime
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Dict, Tuple, Optional

import requests
from requests.auth import HTTPBasicAuth
from requests.adapters import HTTPAdapter
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────── CONFIG ───────────────────────────
BASE_URL   = "https://qualysapi.xx.apps.qualys.in"  # Replace with your actual Qualys URL
FILEPATH   = "C:\\mycode\\CS\\test"                  # Folder containing your EC2 CSV/Excel files
BATCH_SIZE = 50
USERNAME   = "username"    # Replace with real creds
PASSWORD   = "password"    # Replace with real creds
PAGE_SIZE  = 1000
MAX_WORKERS    = 10         # Parallel threads used to fetch the Qualys inventory
PAGE_MAX_RETRIES   = 3      # Retries per page before that page is given up on
PAGE_RETRY_BACKOFF = 1.5    # Seconds, multiplied by attempt number, between retries
CERT_PATH  = True           # Set to path like "C:\\cert.pem" or True to skip verification
ASSET_ID_COL  = "Asset ID"
EC2_RAW_COL   = "sourceInfo.list.Ec2AssetSourceSimple.instanceId"

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)


# ─────────────────────────── AUTH ───────────────────────────
def build_session() -> requests.Session:
    """
    Create a requests.Session with a connection pool large enough for
    MAX_WORKERS concurrent requests. Reused across threads: it carries no
    mutable per-request state once requests pass explicit HTTPBasicAuth.
    """
    session = requests.Session()
    adapter = HTTPAdapter(pool_connections=MAX_WORKERS, pool_maxsize=MAX_WORKERS)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


def login(session: requests.Session) -> None:
    """Login to Qualys and store session cookie."""
    response = session.post(
        f"{BASE_URL}/api/2.0/fo/session/",
        headers={"X-Requested-With": "Python"},
        data={"action": "login", "username": USERNAME, "password": PASSWORD},
        verify=CERT_PATH
    )
    response.raise_for_status()
    if "QualysSession" not in session.cookies:
        raise Exception("Login failed: QualysSession cookie not found")
    log.info("Login successful")


def logout(session: requests.Session) -> None:
    """Logout from Qualys."""
    try:
        session.post(
            f"{BASE_URL}/api/2.0/fo/session/",
            headers={"X-Requested-With": "Python"},
            data={"action": "logout"},
            verify=CERT_PATH
        )
        log.info("Logout successful")
    except Exception:
        pass


# ─────────────────────────── SESSION HELPER ───────────────────────────
def generate_session(session: requests.Session, endpoint: str, body: bytes) -> requests.Response:
    """POST XML request to a Qualys API endpoint."""
    response = session.post(
        f"{BASE_URL}{endpoint}",
        headers={"Content-Type": "application/xml", "Accept": "application/xml"},
        auth=HTTPBasicAuth(USERNAME, PASSWORD),
        data=body,
        verify=CERT_PATH
    )
    response.raise_for_status()
    return response


# ─────────────────────────── VERIFY ASSETS ───────────────────────────
def build_verify_request(asset_ids: List[str]) -> bytes:
    """Build XML body to look up assets by ID."""
    xml = f"""<ServiceRequest>
  <filters>
    <Criteria field="id" operator="IN">
      {",".join(asset_ids)}
    </Criteria>
  </filters>
</ServiceRequest>"""
    return xml.encode("utf-8")


def verify_assets(session: requests.Session, asset_ids: List[str]) -> Dict[str, dict]:
    """Verify which asset IDs exist in Qualys. Returns dict keyed by asset ID."""
    results = {}
    for start in range(0, len(asset_ids), BATCH_SIZE):
        batch = asset_ids[start:start + BATCH_SIZE]
        response = generate_session(session, "/qps/rest/2.0/search/am/hostasset", build_verify_request(batch))
        root = ET.fromstring(response.content)
        assets = root.findall(".//HostAsset")
        for asset in assets:
            asset_id = asset.findtext("id", "").strip()
            hostname = (
                asset.findtext("dnsHostName", "").strip()
                or asset.findtext("netbiosName", "").strip()
            )
            ip = asset.findtext("address", "").strip()
            results[asset_id] = {"found": True, "hostname": hostname, "ip": ip}
        time.sleep(0.2)
    return results


# ─────────────────────────── DELETE / PURGE ───────────────────────────
def build_delete_request(asset_ids: List[str]) -> bytes:
    """Build XML body to delete assets by ID."""
    xml = f"""<ServiceRequest>
  <filters>
    <Criteria field="id" operator="IN">
      {",".join(asset_ids)}
    </Criteria>
  </filters>
</ServiceRequest>"""
    return xml.encode("utf-8")


def purge_batch(session: requests.Session, asset_ids: List[str]) -> List[str]:
    """Purge a single batch of assets. Returns list of deleted IDs."""
    response = generate_session(session, "/qps/rest/2.0/delete/am/hostasset", build_delete_request(asset_ids))
    root = ET.fromstring(response.content)
    response_code = root.findtext(".//responseCode", "")
    if response_code == "SUCCESS":
        return asset_ids
    return []


def purge_assets(session: requests.Session, asset_ids: List[str]) -> List[str]:
    """Purge all assets in batches. Returns list of all deleted IDs."""
    deleted = []
    for start in range(0, len(asset_ids), BATCH_SIZE):
        batch = asset_ids[start:start + BATCH_SIZE]
        deleted.extend(purge_batch(session, batch))
        time.sleep(0.5)
    return deleted


# ─────────────────────────── INPUT FILE READING ───────────────────────────
def get_terminated_instance_dicts(
    filepath: str,
    id_column: str = "Resource ID",
    status_column: str = "State"
) -> List[Dict]:
    """
    Read all CSV/Excel files from filepath, filter rows where status
    contains 'terminated' (stopped instances are intentionally excluded),
    and return as list of dicts.
    """
    files_data = {}
    all_data = pd.DataFrame()

    for filename in os.listdir(filepath):
        file_path = os.path.join(filepath, filename)
        try:
            if filename.endswith(".xlsx") or filename.endswith(".xls"):
                df = pd.read_excel(file_path)
                files_data[filename] = df
                log.info(f"Read Excel file: {filename}")
            elif filename.endswith(".csv"):
                df = pd.read_csv(file_path)
                files_data[filename] = df
                log.info(f"Read CSV file: {filename}")
            else:
                continue
            all_data = pd.concat([all_data, df], ignore_index=True)
        except Exception as e:
            log.error(f"Error reading file {filename}: {str(e)}")

    if all_data.empty:
        log.warning("No data loaded from input files")
        return []

    pattern = r"terminated"
    terminated_df = all_data[all_data[status_column].astype(str).str.contains(pattern, case=False, na=False)]
    result_list = terminated_df[[id_column, status_column]].to_dict(orient="records")
    log.info(f"Total Terminated instances from files: {len(result_list)}")
    return result_list


# ─────────────────────────── FETCH ALL HOSTS ───────────────────────────
def build_request(offset: int) -> bytes:
    """Build paginated XML request for fetching hosts."""
    root = ET.Element("ServiceRequest")
    preferences = ET.SubElement(root, "preferences")
    ET.SubElement(preferences, "startFromOffset").text = str(offset)
    ET.SubElement(preferences, "limitResults").text = str(PAGE_SIZE)
    return ET.tostring(root, encoding="utf-8")


def get_host_count(session: requests.Session) -> int:
    """Ask Qualys for the total number of HostAsset records (no filters)."""
    url = f"{BASE_URL}/qps/rest/2.0/count/am/hostasset"
    auth = HTTPBasicAuth(USERNAME, PASSWORD)
    response = session.post(
        url,
        headers={"Content-Type": "application/xml", "Accept": "application/xml"},
        auth=auth,
        data=b"<ServiceRequest/>",
        verify=CERT_PATH,
        timeout=60
    )
    response.raise_for_status()
    root = ET.fromstring(response.content)

    response_code = root.findtext(".//responseCode", "")
    if response_code and response_code != "SUCCESS":
        raise RuntimeError(f"Count request returned responseCode={response_code!r}")

    count_text = root.findtext(".//count", "")
    if not count_text.isdigit():
        raise RuntimeError(f"Count request returned an unexpected payload: {count_text!r}")
    return int(count_text)


def fetch_host_page(
    session: requests.Session,
    offset: int,
    page_num: int,
    total_pages: Optional[int] = None
) -> List[ET.Element]:
    """Fetch a single offset-bounded page of HostAsset records, with retries."""
    url = f"{BASE_URL}/qps/rest/2.0/search/am/hostasset"
    auth = HTTPBasicAuth(USERNAME, PASSWORD)
    label = f"{page_num}/{total_pages}" if total_pages else str(page_num)
    last_error: Optional[Exception] = None

    for attempt in range(1, PAGE_MAX_RETRIES + 1):
        try:
            response = session.post(
                url,
                headers={"Content-Type": "application/xml", "Accept": "application/xml"},
                auth=auth,
                data=build_request(offset),
                verify=CERT_PATH,
                timeout=60
            )
            response.raise_for_status()
            root = ET.fromstring(response.content)

            response_code = root.findtext(".//responseCode", "")
            if response_code and response_code != "SUCCESS":
                raise RuntimeError(f"responseCode={response_code!r}")

            hosts = root.findall(".//HostAsset")
            log.info(f"Page {label} (offset {offset}-{offset + PAGE_SIZE - 1}) -> {len(hosts)} hosts")
            return hosts
        except Exception as e:
            last_error = e
            log.warning(f"Page {label} (offset {offset}) attempt {attempt}/{PAGE_MAX_RETRIES} failed: {e}")
            if attempt < PAGE_MAX_RETRIES:
                time.sleep(PAGE_RETRY_BACKOFF * attempt)

    log.error(f"Page {label} (offset {offset}) failed after {PAGE_MAX_RETRIES} attempts, giving up")
    raise last_error


def fetch_all_hosts(session: requests.Session, max_workers: int = MAX_WORKERS) -> List[ET.Element]:
    """
    Fetch all HostAsset records from Qualys in parallel.

    Looks up the total record count first, slices that into PAGE_SIZE-sized
    offset ranges (1-1000, 1001-2000, ...), then fetches those ranges
    concurrently across up to `max_workers` threads. Falls back to the
    slower sequential offset-walk if the count lookup itself fails.
    """
    try:
        total_count = get_host_count(session)
    except Exception as e:
        log.warning(f"Count lookup failed ({e}); falling back to sequential pagination")
        return _fetch_all_hosts_sequential(session)

    if total_count <= 0:
        log.info("Qualys reports 0 total host assets")
        return []

    total_pages = math.ceil(total_count / PAGE_SIZE)
    offsets = [1 + i * PAGE_SIZE for i in range(total_pages)]
    log.info(f"Qualys reports {total_count} assets -> {total_pages} pages, fetching with {max_workers} workers")

    all_hosts: List[ET.Element] = []
    failed_pages: List[Tuple[int, int]] = []

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_page = {
            executor.submit(fetch_host_page, session, offset, idx + 1, total_pages): (idx + 1, offset)
            for idx, offset in enumerate(offsets)
        }
        for future in as_completed(future_to_page):
            page_num, offset = future_to_page[future]
            try:
                all_hosts.extend(future.result())
            except Exception as e:
                log.error(f"Giving up on page {page_num} (offset {offset}): {e}")
                failed_pages.append((page_num, offset))

    if failed_pages:
        failed_pages.sort()
        log.warning(f"{len(failed_pages)} page(s) failed permanently: {failed_pages}")

    log.info(f"Total hosts fetched: {len(all_hosts)} (expected {total_count})")
    return all_hosts


def _fetch_all_hosts_sequential(session: requests.Session) -> List[ET.Element]:
    """Original one-page-at-a-time fallback, used only if the count lookup fails."""
    url = f"{BASE_URL}/qps/rest/2.0/search/am/hostasset"
    auth = HTTPBasicAuth(USERNAME, PASSWORD)
    offset = 1
    page = 1
    all_hosts = []

    while True:
        log.info(f"[sequential] Fetching page {page} (offset {offset})")
        response = session.post(
            url,
            headers={"Content-Type": "application/xml", "Accept": "application/xml"},
            auth=auth,
            data=build_request(offset),
            verify=CERT_PATH,
            timeout=60
        )
        response.raise_for_status()

        root = ET.fromstring(response.content)
        hosts = root.findall(".//HostAsset")

        if not hosts:
            break

        all_hosts.extend(hosts)
        has_more = root.findtext(".//hasMoreRecords", "")
        if str(has_more).lower() != "true":
            break

        offset += PAGE_SIZE
        page += 1

    log.info(f"Total hosts fetched (sequential): {len(all_hosts)}")
    return all_hosts


# ─────────────────────────── FLATTEN & DATAFRAME ───────────────────────────
def flatten_element(element: ET.Element, prefix: str = "") -> Dict[str, str]:
    """Recursively flatten an XML element into a dict of col_name -> value."""
    result = {}
    children = list(element)

    if not children:
        value = (element.text or "").strip()
        if prefix:
            result[prefix] = value
        return result

    grouped = defaultdict(list)
    for child in children:
        grouped[child.tag].append(child)

    for tag, siblings in grouped.items():
        column_name = f"{prefix}.{tag}" if prefix else tag

        if len(siblings) == 1:
            child = siblings[0]
            if list(child):
                result.update(flatten_element(child, column_name))
            else:
                result[column_name] = (child.text or "").strip()
        else:
            values = []
            for item in siblings:
                values.append(", ".join(v for v in flatten_element(item).values() if v))
            result[column_name] = " | ".join(values)

    return result


def build_dataframe(hosts: List[ET.Element]) -> pd.DataFrame:
    """Convert list of HostAsset XML elements into a flat DataFrame."""
    rows = []
    for host in hosts:
        rows.append(flatten_element(host))

    df = pd.DataFrame(rows)
    df.fillna("", inplace=True)

    if EC2_RAW_COL in df.columns:
        df.insert(0, "EC2 Instance ID", df[EC2_RAW_COL])
    else:
        df.insert(0, "EC2 Instance ID", "")

    priority_columns = ["id", "address", EC2_RAW_COL]
    front = [col for col in priority_columns if col in df.columns]
    remaining = [col for col in df.columns if col not in front]
    df = df[front + sorted(remaining)]
    return df


# ─────────────────────────── EXPORT ───────────────────────────
def export_excel(df: pd.DataFrame) -> str:
    """Save the full assets DataFrame to Excel."""
    output_file = "qualys_assets.xlsx"
    df.to_excel(output_file, sheet_name="Assets", index=False)
    log.info(f"Saved: {output_file}")
    return output_file


def match_terminated_instances(df: pd.DataFrame, found_ids: List[str]) -> pd.DataFrame:
    """Cross-reference terminated EC2 instance IDs against the Qualys assets DataFrame."""
    if EC2_RAW_COL not in df.columns:
        log.warning(f"Column '{EC2_RAW_COL}' not found in dataframe")
        return pd.DataFrame(columns=["Qualys ID", "address", "EC2 Instance ID"])

    found_set = {str(i).strip() for i in found_ids if str(i).strip()}
    matched = df[df[EC2_RAW_COL].astype(str).str.strip().isin(found_set)]

    result = pd.DataFrame({
        "Qualys ID":      matched["id"] if "id" in matched.columns else "",
        "address":        matched["address"] if "address" in matched.columns else "",
        "EC2 Instance ID": matched[EC2_RAW_COL],
    })
    log.info(f"Matched terminated instances in Qualys: {len(result)}")
    return result


def export_matched_excel(df: pd.DataFrame) -> str:
    """Save the matched terminated-instance DataFrame to Excel."""
    output_file = "qualys_terminated_matches.xlsx"
    df.to_excel(output_file, sheet_name="TerminatedMatches", index=False)
    log.info(f"Saved: {output_file}")
    return output_file


# ─────────────────────────── MAIN ───────────────────────────
def main():
    session = build_session()
    try:
        terminated_instances = get_terminated_instance_dicts(
            FILEPATH,
            id_column="Resource ID",
            status_column="State"
        )
        found_ids = [item["Resource ID"] for item in terminated_instances]

        if not found_ids:
            log.info("No terminated instances found in input files")
            return

        login(session)

        hosts = fetch_all_hosts(session, max_workers=MAX_WORKERS)
        if not hosts:
            log.info("No assets found in Qualys")
            return

        df = build_dataframe(hosts)
        matched_df = match_terminated_instances(df, found_ids)

        if matched_df.empty:
            log.info("No matching assets found in Qualys")
            return

        export_matched_excel(matched_df)

    finally:
        logout(session)


if __name__ == "__main__":
    main()
